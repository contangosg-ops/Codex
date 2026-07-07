from pathlib import Path

from openpyxl import load_workbook


for path in sorted(Path("attachments").glob("*.xls*")):
    print("=" * 80)
    print(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Vol & OI"]
    products = set()
    types = set()
    sample_rows = []

    for row in sheet.iter_rows(values_only=True):
        if row[1]:
            products.add(str(row[1]))
        if row[2]:
            types.add(str(row[2]))
        if row[2] == "Open Interest" and len(sample_rows) < 80:
            sample_rows.append(row[:6])

    print("Products:", sorted(products))
    print("Types:", sorted(types))
    print("Open Interest sample:")
    for row in sample_rows:
        print(row)
    workbook.close()

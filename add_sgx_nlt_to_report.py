import csv
from copy import copy
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import CharacterProperties, Font as DrawingFont, Paragraph, ParagraphProperties, RichTextProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
ATTACHMENTS_DIR = BASE_DIR / "attachments"
REPORT_FILE = BASE_DIR / "Open_Interest_Futures.xlsx"
OUTPUT_FILE = BASE_DIR / "Open_Interest_Futures_with_SGX_EEX.xlsx"
CLEAR_OUTPUT_FILE = BASE_DIR / "Open_Interest_Futures_Combined_Clear.xlsx"
FALLBACK_OUTPUT_FILE = BASE_DIR / "Open_Interest_Futures_Combined_Clear_v2.xlsx"
CSV_FILE = Path.home() / "Downloads" / "myData.csv"

TARGET_CONTRACT = "SGX Baltic Capesize Time Charter Average (5 Routes) 180 Futures"
EEX_PRODUCT = "CPTM"
EEX_SHEET = "EEX Trade Tape"
MONTHS = {
    "07-2026": "Jul-26",
    "08-2026": "Aug-26",
    "Jul 2026": "Jul-26",
    "Aug 2026": "Aug-26",
}
EEX_MONTHS = {
    (2026, 7): "Jul-26",
    (2026, 8): "Aug-26",
}


def latest_sgx_nlt_file():
    files = sorted(
        BASE_DIR.glob("sgx_nlt_*.csv"),
        key=lambda path: datetime.strptime(path.stem.split("_")[-1], "%d%m%y"),
        reverse=True,
    )
    return files[0] if files else CSV_FILE


def row_trade_date(row):
    value = row.get("Traded on (SGT)")
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%d %b %Y %I:%M %p"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def display_date(value):
    return value.strftime("%d %b %Y")


def load_summary():
    summary = defaultdict(lambda: defaultdict(int))
    raw_rows = []
    source_file = latest_sgx_nlt_file()
    source_date = workbook_date_from_name(source_file).date()

    with source_file.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            if TARGET_CONTRACT not in row["Contract"]:
                continue
            trade_date = row_trade_date(row)
            if trade_date is None or trade_date.date() != source_date:
                continue
            month = str(row["Month"]).strip()
            if month not in MONTHS:
                continue
            price = float(str(row["Price"]).replace(",", ""))
            quantity = int(float(str(row["Quantity"]).replace(",", "")))
            summary[MONTHS[month]][price] += quantity
            raw_rows.append(row)

    if len(raw_rows) < 10:
        raise ValueError(
            f"SGX NLT source looks incomplete: only {len(raw_rows)} rows for {source_date}. "
            "Please provide the full SGX export for that traded date."
        )

    return summary, raw_rows, source_file, source_date


def workbook_date_from_name(path):
    stem = path.stem.split("_")[-1]
    return datetime.strptime(stem[:6], "%d%m%y")


def latest_attachment():
    files = sorted(
        ATTACHMENTS_DIR.glob("*.xls*"),
        key=workbook_date_from_name,
        reverse=True,
    )
    if not files:
        raise FileNotFoundError("No Excel attachments found.")
    return files[0]


def load_eex_summary():
    workbook_path = latest_attachment()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[EEX_SHEET]
    headers = list(next(sheet.iter_rows(min_row=4, max_row=4, values_only=True)))
    indexes = {header: index for index, header in enumerate(headers)}

    summary = defaultdict(lambda: defaultdict(int))
    raw_rows = []
    trade_dates = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        trade_date = row[indexes["Date"]]
        contract_month = row[indexes["Contract (Month)"]]
        if not isinstance(contract_month, datetime):
            continue
        month = EEX_MONTHS.get((contract_month.year, contract_month.month))
        if month is None:
            continue
        if row[indexes["Product"]] != EEX_PRODUCT:
            continue
        if row[indexes["Type"]] != "Futures":
            continue

        price = float(row[indexes["Trade Price"]])
        quantity = int(float(row[indexes["Trade Quantity"]]))
        summary[month][price] += quantity
        raw_rows.append(dict(zip(headers, row)))
        if isinstance(trade_date, datetime):
            trade_dates.append(trade_date)

    workbook.close()
    report_date = max(trade_dates).date() if trade_dates else workbook_date_from_name(workbook_path).date()
    return summary, raw_rows, workbook_path, report_date


def combine_summaries(*summaries):
    combined = defaultdict(lambda: defaultdict(int))
    for summary in summaries:
        for month, prices in summary.items():
            for price, quantity in prices.items():
                combined[month][price] += quantity
    return combined


def write_table(sheet, summary, start_row, title):
    dark_red = "C0504D"
    pale_red = "E7B8B7"
    thin_white = Side(style="thin", color="FFFFFF")
    border = Border(left=thin_white, right=thin_white, top=thin_white, bottom=thin_white)

    sheet.cell(start_row, 1, title)
    sheet.cell(start_row, 1).font = Font(bold=True, size=11)

    header_row = start_row + 2
    headers = ["Month", "Price", "Quantity"]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, col, header)
        cell.fill = PatternFill("solid", fgColor=dark_red)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = header_row + 1
    for month in ["Jul-26", "Aug-26"]:
        for price in sorted(summary.get(month, {})):
            sheet.cell(row, 1, month)
            sheet.cell(row, 2, price)
            sheet.cell(row, 3, summary[month][price])
            sheet.cell(row, 2).number_format = "#,##0"
            sheet.cell(row, 3).number_format = "#,##0"
            for col in range(1, 4):
                cell = sheet.cell(row, col)
                cell.alignment = Alignment(horizontal="center")
                if col == 1:
                    cell.fill = PatternFill("solid", fgColor=pale_red)
                    cell.font = Font(bold=True)
            row += 1

    for col in range(1, 4):
        sheet.column_dimensions[get_column_letter(col)].width = max(
            sheet.column_dimensions[get_column_letter(col)].width or 0,
            13,
        )

    return header_row, row - 1


def write_combined_table(sheet, sgx_summary, eex_summary, combined_summary, start_row, title):
    dark_red = "C0504D"
    pale_red = "E7B8B7"
    thin_white = Side(style="thin", color="FFFFFF")
    border = Border(left=thin_white, right=thin_white, top=thin_white, bottom=thin_white)

    sheet.cell(start_row, 1, title)
    sheet.cell(start_row, 1).font = Font(bold=True, size=11)

    header_row = start_row + 2
    headers = ["Month", "Price", "SGX", "EEX", "Total"]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, col, header)
        cell.fill = PatternFill("solid", fgColor=dark_red)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = header_row + 1
    for month in ["Jul-26", "Aug-26"]:
        for price in sorted(combined_summary.get(month, {})):
            sgx_qty = sgx_summary.get(month, {}).get(price, 0)
            eex_qty = eex_summary.get(month, {}).get(price, 0)
            total_qty = combined_summary[month][price]
            values = [month, price, sgx_qty, eex_qty, total_qty]
            for col, value in enumerate(values, 1):
                cell = sheet.cell(row, col, value)
                cell.alignment = Alignment(horizontal="center")
                if col >= 2:
                    cell.number_format = "#,##0"
                if col == 1:
                    cell.fill = PatternFill("solid", fgColor=pale_red)
                    cell.font = Font(bold=True)
            row += 1

    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = max(
            sheet.column_dimensions[get_column_letter(col)].width or 0,
            13,
        )

    return header_row, row - 1


def write_chart_data(sheet, summary, start_col, start_row, month):
    headers = ["Price", "Volume"]
    prices = sorted(summary.get(month, {}))

    for offset, header in enumerate(headers):
        cell = sheet.cell(start_row, start_col + offset, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E7B8B7")
        cell.alignment = Alignment(horizontal="center")

    for row_offset, price in enumerate(prices, 1):
        row = start_row + row_offset
        sheet.cell(row, start_col, price)
        sheet.cell(row, start_col + 1, summary[month][price])
        sheet.cell(row, start_col).number_format = "#,##0"
        sheet.cell(row, start_col + 1).number_format = "#,##0"

    return start_row, start_row + len(prices)


def write_chart_data_for_display(sheet, summary, start_col, start_row, month, min_volume=10):
    headers = ["Price", "Volume"]
    prices = [
        price for price in sorted(summary.get(month, {}))
        if summary[month][price] >= min_volume
    ]

    for offset, header in enumerate(headers):
        cell = sheet.cell(start_row, start_col + offset, header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_offset, price in enumerate(prices, 1):
        row = start_row + row_offset
        sheet.cell(row, start_col, price)
        sheet.cell(row, start_col + 1, summary[month][price])
        sheet.cell(row, start_col).number_format = "#,##0"
        sheet.cell(row, start_col + 1).number_format = "#,##0"

    return start_row, start_row + len(prices)


def add_chart(sheet, min_col, min_row, max_row, anchor, title, y_axis_max=None):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 7
    chart.width = 14
    chart.gapWidth = 90
    chart.visible_cells_only = False
    chart.varyColors = False
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.majorGridlines = chart.y_axis.majorGridlines
    if y_axis_max:
        chart.y_axis.scaling.max = y_axis_max

    data = Reference(sheet, min_col=min_col + 1, max_col=min_col + 1, min_row=min_row, max_row=max_row)
    cats = Reference(sheet, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "4F81BD"
        chart.series[0].graphicalProperties.line.solidFill = "4F81BD"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.dLblPos = "outEnd"
    chart.dataLabels.txPr = dark_chart_text()
    chart.legend = None
    sheet.add_chart(chart, anchor)


def dark_chart_text():
    props = CharacterProperties(
        sz=900,
        b=True,
        solidFill="000000",
        latin=DrawingFont(typeface="Arial"),
    )
    return RichText(
        bodyPr=RichTextProperties(),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=props), endParaRPr=props)],
    )


def add_compact_chart(sheet, min_col, min_row, max_row, anchor, title, y_axis_max=None, source_sheet=None):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 7.2
    chart.width = 16.0
    chart.gapWidth = 95
    chart.visible_cells_only = False
    chart.varyColors = False
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    if y_axis_max:
        chart.y_axis.scaling.max = y_axis_max

    source = source_sheet or sheet
    data = Reference(source, min_col=min_col + 1, max_col=min_col + 1, min_row=min_row, max_row=max_row)
    cats = Reference(source, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "4F81BD"
        chart.series[0].graphicalProperties.line.solidFill = "4F81BD"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.dLblPos = "outEnd"
    chart.dataLabels.txPr = dark_chart_text()
    chart.legend = None
    sheet.add_chart(chart, anchor)


def add_large_chart(sheet, min_col, min_row, max_row, anchor, title, y_axis_max=None, source_sheet=None):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 5.65248
    chart.width = 11.79648
    chart.gapWidth = 95
    chart.visible_cells_only = False
    chart.varyColors = False
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    if y_axis_max:
        chart.y_axis.scaling.max = y_axis_max

    source = source_sheet or sheet
    data = Reference(source, min_col=min_col + 1, max_col=min_col + 1, min_row=min_row, max_row=max_row)
    cats = Reference(source, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "4F81BD"
        chart.series[0].graphicalProperties.line.solidFill = "4F81BD"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.dLblPos = "outEnd"
    chart.dataLabels.txPr = dark_chart_text()
    chart.legend = None
    sheet.add_chart(chart, anchor)


def scale_fonts(sheet, factor=0.8, max_row=60, max_col=12, darken=False):
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue
            font = copy(cell.font)
            if font.sz:
                font.sz = max(6, font.sz * factor)
            else:
                font.sz = 8.8
            font.name = "Arial"
            if darken:
                color = font.color.rgb if font.color is not None and font.color.type == "rgb" else None
                if color not in ("00FFFFFF", "FFFFFFFF", "FFFFFF"):
                    font.color = "FF000000"
            cell.font = font


def darken_numeric_cells(sheet, max_row=80, max_col=12):
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not isinstance(cell.value, (int, float)):
                continue
            font = copy(cell.font)
            font.name = "Microsoft YaHei UI"
            font.bold = True
            font.color = "FF000000"
            cell.font = font


def add_raw_sheet(workbook, sheet_name, headers, raw_rows):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    for row in raw_rows:
        sheet.append([row.get(header, "") for header in headers])
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18
    sheet.column_dimensions["A"].width = 65


def add_trade_summary_sheet(workbook, sheet_name, title, summary, chart_prefix="", y_axis_max=None):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(sheet_name, 0)
    sheet.sheet_view.showGridLines = False

    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)

    write_table(sheet, summary, start_row=3, title=title)
    jul_min_row, jul_max_row = write_chart_data(sheet, summary, start_col=6, start_row=3, month="Jul-26")
    aug_min_row, aug_max_row = write_chart_data(sheet, summary, start_col=9, start_row=3, month="Aug-26")

    add_chart(sheet, min_col=6, min_row=jul_min_row, max_row=jul_max_row, anchor="A24", title=f"{chart_prefix}JUL FFA", y_axis_max=y_axis_max)
    add_chart(sheet, min_col=9, min_row=aug_min_row, max_row=aug_max_row, anchor="A43", title=f"{chart_prefix}AUG FFA", y_axis_max=y_axis_max)

    for col in range(1, 12):
        sheet.column_dimensions[get_column_letter(col)].width = 12


def add_combined_summary_sheet(workbook, sgx_summary, eex_summary, combined_summary, report_date):
    if "Combined Summary" in workbook.sheetnames:
        del workbook["Combined Summary"]
    if "Combined Data" in workbook.sheetnames:
        del workbook["Combined Data"]

    sheet = workbook.create_sheet("Combined Summary", 0)
    data_sheet = workbook.create_sheet("Combined Data", 1)
    sheet.sheet_view.showGridLines = False
    title = f"SGX + EEX Capesize Trades - {display_date(report_date)}"

    source_sheet = workbook["Open Interest Futures"]
    for row in range(1, 12):
        for col in range(1, 11):
            source_cell = source_sheet.cell(row, col)
            target_cell = sheet.cell(row, col)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)

    for col in range(1, 11):
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = source_sheet.column_dimensions[letter].width or 11

    sheet["A13"] = title
    sheet["A13"].font = Font(bold=True, size=11)

    jul_min_row, jul_max_row = write_chart_data_for_display(data_sheet, combined_summary, start_col=1, start_row=2, month="Jul-26")
    aug_min_row, aug_max_row = write_chart_data_for_display(data_sheet, combined_summary, start_col=4, start_row=2, month="Aug-26")
    add_large_chart(sheet, min_col=1, min_row=jul_min_row, max_row=jul_max_row, anchor="A15", title="JUL FFA", y_axis_max=450, source_sheet=data_sheet)
    add_large_chart(sheet, min_col=4, min_row=aug_min_row, max_row=aug_max_row, anchor="A31", title="AUG FFA", y_axis_max=250, source_sheet=data_sheet)

    write_combined_table(data_sheet, sgx_summary, eex_summary, combined_summary, start_row=24, title="Combined Volume")

    for col in range(1, 10):
        sheet.column_dimensions[get_column_letter(col)].width = 8.64

    for col in range(1, 8):
        data_sheet.column_dimensions[get_column_letter(col)].width = 13

    for row in range(1, 60):
        sheet.row_dimensions[row].height = 12
    scale_fonts(sheet, factor=0.6144, max_row=60, max_col=12, darken=True)
    darken_numeric_cells(sheet, max_row=60, max_col=12)

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.print_area = "A1:L49"


def main():
    sgx_summary, sgx_raw_rows, sgx_source_file, sgx_report_date = load_summary()
    eex_summary, eex_raw_rows, eex_source_file, eex_report_date = load_eex_summary()
    report_date = max(sgx_report_date, eex_report_date)
    combined_summary = combine_summaries(sgx_summary, eex_summary)
    workbook = load_workbook(REPORT_FILE)
    sheet = workbook["Open Interest Futures"]
    sheet._charts = []

    write_combined_table(sheet, sgx_summary, eex_summary, combined_summary, start_row=14, title=f"SGX + EEX Capesize Trades - {display_date(report_date)}")
    for row in range(16, 50):
        sheet.row_dimensions[row].height = 13
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 10
    sheet.column_dimensions["A"].width = 9
    sheet.column_dimensions["B"].width = 8
    sheet.column_dimensions["C"].width = 7
    sheet.column_dimensions["D"].width = 7
    sheet.column_dimensions["E"].width = 8
    jul_min_row, jul_max_row = write_chart_data(sheet, combined_summary, start_col=27, start_row=14, month="Jul-26")
    aug_min_row, aug_max_row = write_chart_data(sheet, combined_summary, start_col=30, start_row=14, month="Aug-26")
    add_compact_chart(sheet, min_col=27, min_row=jul_min_row, max_row=jul_max_row, anchor="F13", title="JUL FFA", y_axis_max=450)
    add_compact_chart(sheet, min_col=30, min_row=aug_min_row, max_row=aug_max_row, anchor="F32", title="AUG FFA", y_axis_max=250)
    for col in range(6, 13):
        sheet.column_dimensions[get_column_letter(col)].hidden = False
        sheet.column_dimensions[get_column_letter(col)].width = 11
    for col in range(27, 32):
        sheet.column_dimensions[get_column_letter(col)].hidden = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.print_area = "A1:L49"

    add_trade_summary_sheet(workbook, "EEX Summary", f"EEX CPTM Trades - {display_date(eex_report_date)}", eex_summary, chart_prefix="EEX ", y_axis_max=250)
    add_trade_summary_sheet(workbook, "SGX NLT Summary", f"SGX NLT CWF Trades - {display_date(sgx_report_date)}", sgx_summary, y_axis_max=450)
    add_combined_summary_sheet(workbook, sgx_summary, eex_summary, combined_summary, report_date)
    add_raw_sheet(workbook, "SGX NLT Raw", ["Contract", "Month", "Session", "Price", "Quantity", "Traded on (SGT)", "Cleared On (SGT)"], sgx_raw_rows)
    add_raw_sheet(workbook, "EEX Raw", ["Date", "Time", "Product", "Contract (Month)", "Contract Type", "Contract", "Type", "Options Type", "Options Strike", "Trade Price", "Trade Quantity", "Total Quantity"], eex_raw_rows)

    try:
        workbook.save(REPORT_FILE)
    except PermissionError:
        print(f"Skipped locked file: {REPORT_FILE}")
    try:
        workbook.save(OUTPUT_FILE)
    except PermissionError:
        print(f"Skipped locked file: {OUTPUT_FILE}")
    try:
        workbook.save(CLEAR_OUTPUT_FILE)
        print(CLEAR_OUTPUT_FILE)
    except PermissionError:
        workbook.save(FALLBACK_OUTPUT_FILE)
        print(FALLBACK_OUTPUT_FILE)
    print(f"SGX source: {sgx_source_file}")
    print(f"EEX source: {eex_source_file}")
    print("Summary:")
    for month in ["Jul-26", "Aug-26"]:
        print("Combined", month)
        for price, quantity in sorted(combined_summary.get(month, {}).items()):
            print(f"  {price:,.0f}: {quantity:,}")
        print("SGX", month)
        for price, quantity in sorted(sgx_summary.get(month, {}).items()):
            print(f"  {price:,.0f}: {quantity:,}")
        print("EEX", month)
        for price, quantity in sorted(eex_summary.get(month, {}).items()):
            print(f"  {price:,.0f}: {quantity:,}")


if __name__ == "__main__":
    main()

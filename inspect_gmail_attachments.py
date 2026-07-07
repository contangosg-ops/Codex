import base64
from pathlib import Path

from googleapiclient.errors import HttpError
from openpyxl import load_workbook

from gmail_connect import get_gmail_service


BASE_DIR = Path(__file__).resolve().parent
ATTACHMENTS_DIR = BASE_DIR / "attachments"
ATTACHMENTS_DIR.mkdir(exist_ok=True)


def iter_parts(payload):
    for part in payload.get("parts", []) or []:
        yield part
        yield from iter_parts(part)


def safe_name(name):
    return "".join(ch if ch not in '<>:"/\\|?*' else "_" for ch in name)


def download_excel_attachments(max_messages=20):
    service = get_gmail_service()
    files = []
    seen_messages = set()

    for query in ["filename:xlsx", "filename:xlsm", "filename:xls"]:
        result = service.users().messages().list(
            userId="me",
            q=query,
            maxResults=max_messages,
        ).execute()

        for item in result.get("messages", []):
            message_id = item["id"]
            if message_id in seen_messages:
                continue
            seen_messages.add(message_id)

            message = service.users().messages().get(
                userId="me",
                id=message_id,
                format="full",
            ).execute()
            headers = {
                h.get("name", "").lower(): h.get("value", "")
                for h in message.get("payload", {}).get("headers", [])
            }

            for part in iter_parts(message.get("payload", {})):
                filename = part.get("filename") or ""
                if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
                    continue

                attachment_id = part.get("body", {}).get("attachmentId")
                if not attachment_id:
                    continue

                attachment = service.users().messages().attachments().get(
                    userId="me",
                    messageId=message_id,
                    id=attachment_id,
                ).execute()
                data = base64.urlsafe_b64decode(attachment["data"])
                path = ATTACHMENTS_DIR / f"{message_id}_{safe_name(filename)}"
                path.write_bytes(data)
                files.append(
                    {
                        "path": path,
                        "subject": headers.get("subject", ""),
                        "date": headers.get("date", ""),
                        "filename": filename,
                    }
                )

    return files


def inspect_workbook(path):
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return {"path": str(path), "error": str(exc)}

    info = {"path": str(path), "sheets": workbook.sheetnames}
    target = None
    for sheet_name in workbook.sheetnames:
        normalized = sheet_name.replace(" ", "").lower()
        if normalized == "vol&oi":
            target = sheet_name
            break

    if target:
        sheet = workbook[target]
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True):
            rows.append([value for value in row[:20]])
        info["vol_oi_sheet"] = target
        info["preview"] = rows

    workbook.close()
    return info


def main():
    try:
        files = download_excel_attachments()
    except HttpError as exc:
        print(f"Gmail API error: {exc}")
        return

    if not files:
        print("No Excel attachments found.")
        return

    for item in files:
        print("=" * 80)
        print(f"Subject: {item['subject']}")
        print(f"Date: {item['date']}")
        print(f"Attachment: {item['filename']}")
        info = inspect_workbook(item["path"])
        print(f"Saved: {info['path']}")
        if "error" in info:
            print(f"Error: {info['error']}")
            continue
        print(f"Sheets: {', '.join(info['sheets'])}")
        if "vol_oi_sheet" in info:
            print(f"Vol&OI sheet: {info['vol_oi_sheet']}")
            for row in info["preview"]:
                print(row)


if __name__ == "__main__":
    main()

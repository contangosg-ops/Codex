from collections import defaultdict
import concurrent.futures
import csv
from datetime import datetime, timedelta
import io
import json
import time
from pathlib import Path
import urllib.parse
import urllib.error
import urllib.request
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
ATTACHMENTS_DIR = BASE_DIR / "attachments"
OUTPUT_FILE = BASE_DIR / "Open_Interest_Futures.xlsx"
CACHE_DIR = BASE_DIR / ".cache"

PRODUCTS = {
    "Capes": {"eex": ("CPTM", "Capesize", "5TC"), "sgx": "CWF"},
    "Pmax": {"eex": ("P5TC", "Panamax", "5TC"), "sgx": "PWF"},
    "Smax": {"eex": ("SPTM", "Supramax", "10TC"), "sgx": "SWF"},
}

EEX_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Origin": "https://www.eex.com",
    "Referer": "https://www.eex.com/en/market-data/market-data-hub",
    "X-Requested-With": "XMLHttpRequest",
}

SGX_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "*/*",
    "Referer": "https://www.sgx.com/",
}

SGX_DAILY_URL = "https://links.sgx.com/1.0.0/derivatives-daily/{key}/FUTURE.zip"
SGX_KEY_ANCHOR = (datetime(2026, 7, 3), 7547)

ROWS = [
    ("Jul-26", [datetime(2026, 7, 1)]),
    ("Aug-26", [datetime(2026, 8, 1)]),
    ("Sep-26", [datetime(2026, 9, 1)]),
    ("Q3 26", [datetime(2026, month, 1) for month in (7, 8, 9)]),
    ("Q4 26", [datetime(2026, month, 1) for month in (10, 11, 12)]),
    ("Q1 27", [datetime(2027, month, 1) for month in (1, 2, 3)]),
    ("Cal2027", [datetime(2027, month, 1) for month in range(1, 13)]),
    ("Cal2028", [datetime(2028, month, 1) for month in range(1, 13)]),
]


def workbook_date_from_name(path):
    stem = path.stem.split("_")[-1]
    return datetime.strptime(stem[:6], "%d%m%y")


def latest_two_workbooks():
    files = sorted(
        ATTACHMENTS_DIR.glob("*.xls*"),
        key=workbook_date_from_name,
        reverse=True,
    )
    if len(files) < 2:
        raise FileNotFoundError("Need at least two Excel attachments in the attachments folder.")
    return files[:2]


def fetch_json(url, headers):
    CACHE_DIR.mkdir(exist_ok=True)
    cache_key = str(abs(hash(url)))
    cache_file = CACHE_DIR / f"{cache_key}.json"
    if cache_file.exists():
        return json.loads(cache_file.read_text(encoding="utf-8"))

    request = urllib.request.Request(url, headers=headers)
    last_error = None
    for attempt in range(6):
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                payload = json.load(response)
                cache_file.write_text(json.dumps(payload), encoding="utf-8")
                return payload
        except urllib.error.HTTPError as error:
            last_error = error
            if error.code != 429:
                raise
            time.sleep(10 + attempt * 10)
    raise last_error


def fetch_eex_open_interest(trade_date, months):
    data = defaultdict(dict)

    def fetch_one(product_name, config, month):
        short_code, product, product_specific = config["eex"]
        params = {
            "shortCode": short_code,
            "commodity": "FREIGHT",
            "pricing": "F",
            "area": "Freight",
            "product": product,
            "productSpecific": product_specific,
            "maturity": month.strftime("%Y%m"),
            "maturityType": "Month",
            "startDate": trade_date.strftime("%Y-%m-%d"),
            "endDate": trade_date.strftime("%Y-%m-%d"),
            "isRolling": "true",
        }
        url = (
            "https://api.eex-group.com/pub/market-data/table-data?"
            + urllib.parse.urlencode(params)
        )
        payload = fetch_json(url, EEX_HEADERS)
        indexes = {header: index for index, header in enumerate(payload["header"])}
        value = 0
        for row in payload.get("data", []):
            if row[indexes["tradeDate"]] == trade_date.strftime("%Y-%m-%d"):
                value = int(row[indexes["grossOpenInt"]] or 0)
                break
        return product_name, month, value

    for product_name, config in PRODUCTS.items():
        for month in months:
            product_name, month, value = fetch_one(product_name, config, month)
            data[product_name][month] = value
            time.sleep(0.4)

    return data


def read_sgx_future_zip(key):
    request = urllib.request.Request(SGX_DAILY_URL.format(key=key), headers=SGX_HEADERS)
    with urllib.request.urlopen(request, timeout=30) as response:
        raw = response.read()
    if not raw.startswith(b"PK"):
        raise ValueError(f"SGX daily file {key} is not a zip file.")

    archive = zipfile.ZipFile(io.BytesIO(raw))
    csv_name = archive.namelist()[0]
    rows = list(csv.DictReader(io.TextIOWrapper(archive.open(csv_name), encoding="utf-8")))
    if not rows:
        raise ValueError(f"SGX daily file {key} has no rows.")
    file_date = datetime.strptime(rows[0]["DATE"], "%Y%m%d")
    return file_date, rows


def find_sgx_key(trade_date):
    anchor_date, anchor_key = SGX_KEY_ANCHOR
    rough_key = anchor_key + (trade_date - anchor_date).days
    for key in range(rough_key - 20, rough_key + 21):
        try:
            file_date, _rows = read_sgx_future_zip(key)
        except Exception:
            continue
        if file_date.date() == trade_date.date():
            return key
    raise FileNotFoundError(f"Could not find SGX FUTURE.zip for {trade_date:%Y-%m-%d}.")


def latest_two_available_trade_dates(reference_date=None):
    if reference_date is None:
        reference_date = datetime.now()

    trade_dates = []
    check_date = reference_date - timedelta(days=1)
    while len(trade_dates) < 2:
        try:
            find_sgx_key(check_date)
        except FileNotFoundError:
            check_date -= timedelta(days=1)
            continue
        trade_dates.append(check_date)
        check_date -= timedelta(days=1)

    return trade_dates


def fetch_sgx_open_interest(trade_date):
    key = find_sgx_key(trade_date)
    _file_date, rows = read_sgx_future_zip(key)
    data = defaultdict(dict)

    product_codes = {config["sgx"]: product_name for product_name, config in PRODUCTS.items()}
    for row in rows:
        product_name = product_codes.get(row["COM"].strip())
        if not product_name:
            continue
        month = datetime(int(row["COM_YY"]), int(row["COM_MM"]), 1)
        data[product_name][month] = int(float(row["OINT"] or 0))

    return data, key


def load_open_interest_from_web(trade_date, months):
    eex_data = fetch_eex_open_interest(trade_date, months)
    sgx_data, sgx_key = fetch_sgx_open_interest(trade_date)
    combined = defaultdict(dict)

    for product_name in PRODUCTS:
        for month in months:
            combined[product_name][month] = (
                eex_data.get(product_name, {}).get(month, 0)
                + sgx_data.get(product_name, {}).get(month, 0)
            )

    return combined, {"sgx_key": sgx_key, "eex": eex_data, "sgx": sgx_data}


def value_for(data, product, months):
    return sum(data.get(product, {}).get(month, 0) for month in months)


def short_date(value):
    return f"{value.day}-{value.strftime('%b')}"


def write_report():
    current_date, previous_date = latest_two_available_trade_dates()
    required_months = sorted({month for _label, months in ROWS for month in months})
    current_data, current_source = load_open_interest_from_web(current_date, required_months)
    previous_data, previous_source = load_open_interest_from_web(previous_date, required_months)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Open Interest Futures"

    dark_red = "C0504D"
    pale_red = "E7B8B7"
    green = "63BE7B"
    yellow = "FFEB84"
    red = "F8696B"
    thin_white = Side(style="thin", color="FFFFFF")
    border = Border(left=thin_white, right=thin_white, top=thin_white, bottom=thin_white)

    sheet["A1"] = "Open Interest Futures"
    sheet["A1"].font = Font(bold=True, size=11)

    start_col = 2
    for group_index, product_name in enumerate(PRODUCTS):
        col = start_col + group_index * 3
        sheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        cell = sheet.cell(2, col, product_name)
        cell.fill = PatternFill("solid", fgColor=dark_red)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

        for offset, header in enumerate(
            [
                short_date(current_date),
                short_date(previous_date),
                "Change",
            ]
        ):
            header_cell = sheet.cell(3, col + offset, header)
            header_cell.fill = PatternFill("solid", fgColor=pale_red)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal="center")

    for row_index, (label, months) in enumerate(ROWS, start=4):
        label_cell = sheet.cell(row_index, 1, label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")

        for group_index, product_name in enumerate(PRODUCTS):
            col = start_col + group_index * 3
            current_value = value_for(current_data, product_name, months)
            previous_value = value_for(previous_data, product_name, months)
            change = current_value - previous_value

            values = [current_value, previous_value, change]
            for offset, value in enumerate(values):
                cell = sheet.cell(row_index, col + offset, value)
                cell.number_format = "#,##0;[Red]-#,##0;0"
                cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=10):
        for cell in row:
            cell.border = border

    for col in range(1, 11):
        sheet.column_dimensions[get_column_letter(col)].width = 11
    sheet.column_dimensions["A"].width = 12

    for change_col in ("D", "G", "J"):
        sheet.conditional_formatting.add(
            f"{change_col}4:{change_col}11",
            ColorScaleRule(
                start_type="min",
                start_color=red,
                mid_type="num",
                mid_value=0,
                mid_color=yellow,
                end_type="max",
                end_color=green,
            ),
        )

    sheet.freeze_panes = "B4"

    source = workbook.create_sheet("Source")
    source.append(["Current source", "EEX Market Data Hub + SGX FUTURE.zip"])
    source.append(["Previous source", "EEX Market Data Hub + SGX FUTURE.zip"])
    source.append(["Current date", current_date.strftime("%d-%b-%Y")])
    source.append(["Previous date", previous_date.strftime("%d-%b-%Y")])
    source.append(["Current SGX key", current_source["sgx_key"]])
    source.append(["Previous SGX key", previous_source["sgx_key"]])
    source.append([])
    source.append(["Row", "Product", "Current total", "Previous total", "Change"])
    for label, months in ROWS:
        for product_name in PRODUCTS:
            current_value = value_for(current_data, product_name, months)
            previous_value = value_for(previous_data, product_name, months)
            source.append([label, product_name, current_value, previous_value, current_value - previous_value])

    workbook.save(OUTPUT_FILE)
    return OUTPUT_FILE


if __name__ == "__main__":
    print(write_report())
